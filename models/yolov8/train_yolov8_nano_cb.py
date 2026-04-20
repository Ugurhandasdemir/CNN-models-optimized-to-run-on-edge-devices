"""
train_yolov8_cb.py
==================
YOLOv8-nano eğitim scripti — Class-Balanced Loss + Excel metrik kaydı.

KURULUM:
    pip install ultralytics openpyxl

KULLANIM:
    python train_yolov8_cb.py \
        --data   data.yaml \
        --counts 50000 10000 3000 800

    Tüm eğitim parametreleri NANO_CFG üzerinden yönetilir.
    CLI argümanları NANO_CFG değerlerini override eder.

NOTLAR:
    - cb_loss.py bu scriptile aynı dizinde olmalıdır.
    - CB ağırlıkları model.class_weights üzerinden v8DetectionLoss'a
      inject edilir; box ve dfl loss değişmez.
    - Her epoch sonunda Excel'e bir satır yazılır.

CB LOSS ENTEGRASYON MEKANİZMASI:
    Ultralytics'in v8DetectionLoss sınıfı, criterion init sırasında
    model.class_weights attribute'unu otomatik okur:

        self.class_weights = getattr(model, "class_weights", None)
        if self.class_weights is not None:
            bce_loss *= self.class_weights   # (bs, anchors, nc)

    on_pretrain_routine_end callback'inde model bu attribute'a set
    edilir; criterion lazy init olduğundan bu zamanlama kritiktir.
"""

import argparse
from pathlib import Path
from typing import Optional

import torch
import torch.nn as nn
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from ultralytics import YOLO
from ultralytics.utils import LOGGER

from cb_loss import compute_cb_weights


# ──────────────────────────────────────────────────────────────
# 0. Eğitim konfigürasyonu — YOLOv8-nano varsayılanları
# ──────────────────────────────────────────────────────────────

NANO_CFG = {
    # ── Temel ─────────────────────────────────────────────────
    "IMG_SIZE":       640,
    "DEVICE":         0,          # GPU index; "cpu" de olabilir
    "EPOCHS":         1000,
    "BATCH_SIZE":     32,
    "PATIENCE":       30,         # Early stopping; 0 → kapalı
    "WORKERS":        4,
    "CACHE":          False,      # True / "ram" / "disk"

    # ── Optimizer ─────────────────────────────────────────────
    "OPTIMIZER":      "SGD",      # "SGD" | "Adam" | "AdamW"
    "LR0":            0.01,       # Başlangıç öğrenme oranı
    "LRF":            0.1,       # Son LR = LR0 * LRF  (cosine decay)
    "MOMENTUM":       0.937,      # SGD momentum / Adam beta1
    "WEIGHT_DECAY":   0.0005,
    "WARMUP_EPOCHS":  10,         # LR warmup süresi (epoch)
    "AMP":            True,       # Automatic Mixed Precision

    # ── Augmentation ──────────────────────────────────────────
    "HSV_H":          0.015,      # Hue jitter
    "HSV_S":          0.7,        # Saturation jitter
    "HSV_V":          0.4,        # Value jitter
    "MOSAIC":         1.0,        # Mosaic olasılığı (0–1)
    "MIXUP":          0.0,        # MixUp olasılığı  (0–1)
    "FLIPUD":         0.0,        # Dikey flip olasılığı
    "FLIPLR":         0.0,        # Yatay flip olasılığı
    "DEGREES":        0.0,        # Rotasyon aralığı (±derece)
    "TRANSLATE":      0.0,        # Öteleme (görüntü boyutunun oranı)
    "SCALE":          0.0,        # Ölçek jitter (±oran)
    "PERSPECTIVE":    0.0,        # Perspektif bozulması (0–0.001)
    "ERASING":        0.0,        # Random erasing olasılığı

    # ── CB Loss ───────────────────────────────────────────────
    # gamma yok: v8DetectionLoss BCE tabanlıdır, focal gamma
    # kullanmaz. Sınıf dengesizliği CB ağırlıklarıyla (beta)
    # ele alınır.
    "BETA":        0.99999,

    # ── Çıktı ─────────────────────────────────────────────────
    "PROJECT":        "",
    "NAME":           "yolov8n_cb_2",
    "OUTPUT_XLSX":    "training_results_2.xlsx",
}


# ──────────────────────────────────────────────────────────────
# 1. CB Loss entegrasyonu — model.class_weights injection
# ──────────────────────────────────────────────────────────────

def inject_cb_weights(
    model,
    class_counts: list[int],
    beta: float,
    device: torch.device,
) -> None:
    """
    CB ağırlıklarını model.class_weights olarak set eder.

    v8DetectionLoss, __init__ sırasında aşağıdaki satırla bu
    attribute'u otomatik okur ve her BCE hesabına uygular:

        self.class_weights = getattr(model, "class_weights", None)
        if self.class_weights is not None:
            bce_loss *= self.class_weights   # shape (1, 1, nc)

    Parametreler
    ------------
    model        : DDP unwrap edilmiş YOLO DetectionModel
    class_counts : her foreground sınıfına ait örnek sayıları
    beta         : CB beta hiperparametresi (0 < β < 1)
    device       : modelin bulunduğu cihaz
    """
    weights = compute_cb_weights(
        class_counts=class_counts,
        beta=beta,
        normalize=True,
        device=device,
    )
    # bce_loss shape'i (bs, num_anchors, nc) olduğundan
    # broadcasting için (1, 1, nc) gereklidir.
    model.class_weights = weights.view(1, 1, -1)

    LOGGER.info(
        f"[CB-Train] class_weights inject edildi — "
        f"beta={beta}, nc={len(class_counts)}, "
        f"min={weights.min():.4f}, max={weights.max():.4f}"
    )


# ──────────────────────────────────────────────────────────────
# 2. Excel şablonu oluştur / yükle
# ──────────────────────────────────────────────────────────────

HEADER_ROW_1 = [
    "Epoch", "Learning Rate",
    "Box Loss", "",
    "Class Loss", "",
    "DFL Loss", "",
    "Precision", "",
    "Recall", "",
    "F1 Score", "",
    "mAP", "",
    "Accuracy (val)",
]

HEADER_ROW_2 = [
    "", "",
    "Train", "Val",
    "Train", "Val",
    "Train", "Val",
    "Val", "Train*",
    "Val", "Train*",
    "Val", "Train*",
    "mAP@0.5", "mAP@0.5:0.95",
    "",
]

_THIN        = Side(border_style="thin", color="BBBBBB")
BORDER       = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
HEADER_FILL  = PatternFill("solid", start_color="1F3864")
SUBHEAD_FILL = PatternFill("solid", start_color="2E75B6")
HEADER_FONT  = Font(name="Arial", bold=True, color="FFFFFF", size=10)
DATA_FONT    = Font(name="Arial", size=10)
COL_WIDTHS   = [7, 14, 10, 10, 10, 10, 10, 10, 10, 10, 10, 10, 10, 10, 10, 12, 14]


def _apply_header(ws) -> None:
    """Training Metrics sheet'ine çift satır header uygular."""
    # Grup başlıklarını merge et (Box Loss, Class Loss, ...)
    for start, end in [(3,4),(5,6),(7,8),(9,10),(11,12),(13,14),(15,16)]:
        ws.merge_cells(start_row=1, start_column=start,
                       end_row=1, end_column=end)
    # Epoch, LR ve Accuracy sütunları her iki satırı kapsar
    for col in (1, 2, 17):
        ws.merge_cells(start_row=1, start_column=col,
                       end_row=2, end_column=col)

    for ci, val in enumerate(HEADER_ROW_1, 1):
        c = ws.cell(row=1, column=ci, value=val or None)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = Alignment(horizontal="center", vertical="center",
                                wrap_text=True)
        c.border = BORDER

    for ci, val in enumerate(HEADER_ROW_2, 1):
        c = ws.cell(row=2, column=ci, value=val or None)
        c.fill = SUBHEAD_FILL
        c.font = HEADER_FONT
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = BORDER

    for i, w in enumerate(COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 22
    ws.freeze_panes = "A3"


def _write_note(ws, row: int) -> None:
    """Dipnotu belirtilen satıra yazar."""
    note = (
        "* Train Precision/Recall/F1: YOLO varsayılan olarak sadece val setinde "
        "raporlar. Modeliniz train değeri üretiyorsa doldurunuz."
    )
    c = ws.cell(row=row, column=1, value=note)
    c.font = Font(name="Arial", size=9, italic=True, color="666666")
    ws.merge_cells(start_row=row, start_column=1,
                   end_row=row, end_column=17)


def build_excel(output_path: Path) -> None:
    """Sıfırdan üç sheet'li Excel dosyası oluşturur."""
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()

    ws = wb.active
    ws.title = "Training Metrics"
    _apply_header(ws)
    _write_note(ws, row=3)

    cm = wb.create_sheet("Confusion Matrix")
    cm["A1"] = "Confusion Matrix — Eğitim tamamlandıktan sonra doldurulur."
    cm["A1"].font = Font(name="Arial", bold=True, size=11)

    ch = wb.create_sheet("Charts")
    for r, txt in enumerate([
        "Grafikler bu sheet'e eklenecektir.",
        "Insert → Chart ile Training Metrics verilerini kullanabilirsiniz.",
    ], start=1):
        c = ch.cell(row=r, column=1, value=txt)
        c.font = Font(name="Arial", size=10, italic=True, color="555555")

    wb.save(output_path)
    LOGGER.info(f"[CB-Train] Excel oluşturuldu: {output_path}")


def _fmt(v) -> Optional[float]:
    """None'ı korur, float değeri 4 ondalığa yuvarlar."""
    return None if v is None else round(float(v), 4)


def append_epoch_row(
    output_path: Path,
    epoch: int,
    lr: float,
    box_loss_train: Optional[float],
    box_loss_val:   Optional[float],
    cls_loss_train: Optional[float],
    cls_loss_val:   Optional[float],
    dfl_loss_train: Optional[float],
    dfl_loss_val:   Optional[float],
    precision_val:  Optional[float],
    recall_val:     Optional[float],
    f1_val:         Optional[float],
    map50:          Optional[float],
    map50_95:       Optional[float],
    accuracy_val:   Optional[float],
) -> None:
    """
    Mevcut Excel'e bir epoch satırı ekler.

    Strateji: notu sil → veri ekle → notu en alta geri yaz.
    Bu sayede dipnot her zaman son satırda kalır.
    """
    wb = openpyxl.load_workbook(output_path)
    ws = wb["Training Metrics"]

    # Dipnotu bul ve sil
    note_row = None
    for row in ws.iter_rows(min_row=3):
        for cell in row:
            if cell.value and "Train Precision" in str(cell.value):
                note_row = cell.row
                break
        if note_row:
            break
    if note_row:
        ws.delete_rows(note_row)

    data_row = ws.max_row + 1

    row_data = [
        epoch, round(lr, 8),
        _fmt(box_loss_train), _fmt(box_loss_val),
        _fmt(cls_loss_train), _fmt(cls_loss_val),
        _fmt(dfl_loss_train), _fmt(dfl_loss_val),
        _fmt(precision_val),  None,   # Train* precision — boş
        _fmt(recall_val),     None,   # Train* recall    — boş
        _fmt(f1_val),         None,   # Train* F1        — boş
        _fmt(map50), _fmt(map50_95),
        _fmt(accuracy_val),
    ]

    alt_fill = (PatternFill("solid", start_color="EBF3FB")
                if data_row % 2 == 0 else None)

    for ci, val in enumerate(row_data, 1):
        c = ws.cell(row=data_row, column=ci, value=val)
        c.font = DATA_FONT
        c.alignment = Alignment(horizontal="center")
        c.border = BORDER
        if alt_fill:
            c.fill = alt_fill

    _write_note(ws, row=data_row + 1)
    wb.save(output_path)


# ──────────────────────────────────────────────────────────────
# 3. Trainer callback'leri
# ──────────────────────────────────────────────────────────────

class CBTrainingCallbacks:
    """
    YOLO trainer'ına eklenen callback sınıfı.

    Hook zamanlaması:
      on_pretrain_routine_end  →  CB ağırlıklarını modele inject et.
          Model device'a taşınmış, criterion henüz init edilmemiş.
          v8DetectionLoss lazy init sırasında model.class_weights'i okur.
      on_fit_epoch_end         →  metrikleri topla, Excel'e yaz.
      on_train_end             →  tamamlanma logu.
    """

    def __init__(
        self,
        output_path:  Path,
        class_counts: list[int],
        beta:         float,
    ):
        self.output_path  = output_path
        self.class_counts = class_counts
        self.beta         = beta

        if not output_path.exists():
            build_excel(output_path)

    # ── hook 1 ────────────────────────────────────────────────

    def on_pretrain_routine_end(self, trainer) -> None:
        """
        CB ağırlıklarını inject et.

        _setup_train'in sonunda tetiklenir:
          ✓ model GPU'ya taşınmış
          ✓ criterion henüz oluşturulmamış   ← kritik pencere
          ✓ v8DetectionLoss ilk loss() çağrısında lazy init olacak
            ve o sırada model.class_weights'i okuyacak
        """
        # DDP sarmalıysa iç modeli al
        raw = trainer.model
        while isinstance(raw, nn.parallel.DistributedDataParallel):
            raw = raw.module

        device = next(raw.parameters()).device
        inject_cb_weights(raw, self.class_counts, self.beta, device)

    # ── hook 2 ────────────────────────────────────────────────

    def on_fit_epoch_end(self, trainer) -> None:
        """Metrikleri topla, Excel'e yaz."""
        epoch      = trainer.epoch
        metrics    = trainer.metrics
        loss_items = trainer.loss_items
        lr_list    = trainer.scheduler.get_last_lr()
        lr_val     = lr_list[0] if lr_list else trainer.args.lr0

        # Train loss üçlüsü: [box, cls, dfl]
        def _li(i):
            return float(loss_items[i]) if loss_items is not None else None

        box_t, cls_t, dfl_t = _li(0), _li(1), _li(2)

        # Val metrikleri — key'ler Ultralytics sürümüne bağlı
        def _m(key):
            return float(metrics[key]) if key in metrics else None

        box_v    = _m("val/box_loss")
        cls_v    = _m("val/cls_loss")
        dfl_v    = _m("val/dfl_loss")
        prec     = _m("metrics/precision(B)")
        rec      = _m("metrics/recall(B)")
        map50    = _m("metrics/mAP50(B)")
        map50_95 = _m("metrics/mAP50-95(B)")
        accuracy = _m("metrics/accuracy_top1")

        f1 = None
        if prec is not None and rec is not None:
            denom = prec + rec
            f1 = (2 * prec * rec / denom) if denom > 0 else 0.0

        append_epoch_row(
            output_path=self.output_path,
            epoch=epoch, lr=lr_val,
            box_loss_train=box_t, box_loss_val=box_v,
            cls_loss_train=cls_t, cls_loss_val=cls_v,
            dfl_loss_train=dfl_t, dfl_loss_val=dfl_v,
            precision_val=prec, recall_val=rec, f1_val=f1,
            map50=map50, map50_95=map50_95,
            accuracy_val=accuracy,
        )

        if all(v is not None for v in [box_t, cls_t, dfl_t, map50]):
            LOGGER.info(
                f"[CB-Train] Epoch {epoch:3d} → "
                f"box={box_t:.4f}  cls={cls_t:.4f}  dfl={dfl_t:.4f}  "
                f"mAP@0.5={map50:.4f}"
            )
        else:
            LOGGER.info(f"[CB-Train] Epoch {epoch} kaydedildi.")

    # ── hook 3 ────────────────────────────────────────────────

    def on_train_end(self, trainer) -> None:
        LOGGER.info(
            f"[CB-Train] Eğitim tamamlandı. Sonuçlar: {self.output_path}"
        )


# ──────────────────────────────────────────────────────────────
# 4. Ana eğitim fonksiyonu
# ──────────────────────────────────────────────────────────────

def train(
    data:         str,
    class_counts: list[int],
    cfg:          Optional[dict] = None,
) -> None:
    """
    YOLOv8-nano eğitimini başlatır.

    cfg verilmezse NANO_CFG kullanılır.
    cfg verilirse NANO_CFG üzerine override edilir;
    sadece değiştirmek istediğin anahtarları geçmen yeterli.
    """
    c = {**NANO_CFG, **(cfg or {})}

    import importlib.util
    if importlib.util.find_spec("cb_loss") is None:
        raise ImportError(
            "cb_loss modülü bulunamadı. "
            "cb_loss.py'nin çalışma dizininde ya da PYTHONPATH'te "
            "olduğundan emin olun."
        )

    model = YOLO("yolov8n.pt")

    cbs = CBTrainingCallbacks(
        output_path=Path(c["OUTPUT_XLSX"]),
        class_counts=class_counts,
        beta=c["BETA"],
    )

    model.add_callback("on_pretrain_routine_end", cbs.on_pretrain_routine_end)
    model.add_callback("on_fit_epoch_end",        cbs.on_fit_epoch_end)
    model.add_callback("on_train_end",            cbs.on_train_end)

    model.train(
        data=data,
        # ── Temel ─────────────────────────────────────────────
        imgsz=c["IMG_SIZE"],
        device=c["DEVICE"],
        epochs=c["EPOCHS"],
        batch=c["BATCH_SIZE"],
        patience=c["PATIENCE"],
        workers=c["WORKERS"],
        cache=c["CACHE"],
        # ── Optimizer ─────────────────────────────────────────
        optimizer=c["OPTIMIZER"],
        lr0=c["LR0"],
        lrf=c["LRF"],
        momentum=c["MOMENTUM"],
        weight_decay=c["WEIGHT_DECAY"],
        warmup_epochs=c["WARMUP_EPOCHS"],
        amp=c["AMP"],
        # ── Augmentation ──────────────────────────────────────
        hsv_h=c["HSV_H"],
        hsv_s=c["HSV_S"],
        hsv_v=c["HSV_V"],
        mosaic=c["MOSAIC"],
        mixup=c["MIXUP"],
        flipud=c["FLIPUD"],
        fliplr=c["FLIPLR"],
        degrees=c["DEGREES"],
        translate=c["TRANSLATE"],
        scale=c["SCALE"],
        perspective=c["PERSPECTIVE"],
        erasing=c["ERASING"],
        # ── Çıktı ─────────────────────────────────────────────
        project=c["PROJECT"],
        name=c["NAME"],
        verbose=True,
    )


# ──────────────────────────────────────────────────────────────
# 5. CLI
# ──────────────────────────────────────────────────────────────

def parse_args():
    p = argparse.ArgumentParser(
        description="YOLOv8-nano + Class-Balanced Loss eğitimi",
        formatter_class=argparse.ArgumentDefaultsHelpFormatter,
    )

    # Zorunlu
    p.add_argument("--data",   required=True,
                   help="data.yaml yolu")
    p.add_argument("--counts", type=int, nargs="+", required=True,
                   help="Her sınıf için örnek sayısı (sınıf sırasıyla)")

    # NANO_CFG override'ları
    p.add_argument("--imgsz",         type=int,   default=NANO_CFG["IMG_SIZE"])
    p.add_argument("--device",                    default=str(NANO_CFG["DEVICE"]))
    p.add_argument("--epochs",        type=int,   default=NANO_CFG["EPOCHS"])
    p.add_argument("--batch",         type=int,   default=NANO_CFG["BATCH_SIZE"])
    p.add_argument("--patience",      type=int,   default=NANO_CFG["PATIENCE"])
    p.add_argument("--workers",       type=int,   default=NANO_CFG["WORKERS"])
    p.add_argument("--cache",                     default=str(NANO_CFG["CACHE"]))

    p.add_argument("--optimizer",                 default=NANO_CFG["OPTIMIZER"])
    p.add_argument("--lr0",           type=float, default=NANO_CFG["LR0"])
    p.add_argument("--lrf",           type=float, default=NANO_CFG["LRF"])
    p.add_argument("--momentum",      type=float, default=NANO_CFG["MOMENTUM"])
    p.add_argument("--weight-decay",  type=float, default=NANO_CFG["WEIGHT_DECAY"],
                   dest="weight_decay")
    p.add_argument("--warmup-epochs", type=int,   default=NANO_CFG["WARMUP_EPOCHS"],
                   dest="warmup_epochs")
    p.add_argument("--no-amp",        action="store_false", dest="amp")
    p.set_defaults(amp=NANO_CFG["AMP"])

    p.add_argument("--hsv-h",       type=float, default=NANO_CFG["HSV_H"],       dest="hsv_h")
    p.add_argument("--hsv-s",       type=float, default=NANO_CFG["HSV_S"],       dest="hsv_s")
    p.add_argument("--hsv-v",       type=float, default=NANO_CFG["HSV_V"],       dest="hsv_v")
    p.add_argument("--mosaic",      type=float, default=NANO_CFG["MOSAIC"])
    p.add_argument("--mixup",       type=float, default=NANO_CFG["MIXUP"])
    p.add_argument("--flipud",      type=float, default=NANO_CFG["FLIPUD"])
    p.add_argument("--fliplr",      type=float, default=NANO_CFG["FLIPLR"])
    p.add_argument("--degrees",     type=float, default=NANO_CFG["DEGREES"])
    p.add_argument("--translate",   type=float, default=NANO_CFG["TRANSLATE"])
    p.add_argument("--scale",       type=float, default=NANO_CFG["SCALE"])
    p.add_argument("--perspective", type=float, default=NANO_CFG["PERSPECTIVE"])
    p.add_argument("--erasing",     type=float, default=NANO_CFG["ERASING"])

    p.add_argument("--beta",    type=float, default=NANO_CFG["BETA"],
                   help="CB Loss beta (0 < β < 1)")
    p.add_argument("--project",             default=NANO_CFG["PROJECT"])
    p.add_argument("--name",                default=NANO_CFG["NAME"])
    p.add_argument("--output",              default=NANO_CFG["OUTPUT_XLSX"],
                   help="Excel çıktı dosyası")

    return p.parse_args()


if __name__ == "__main__":
    args = parse_args()

    override = {
        "IMG_SIZE":      args.imgsz,
        "DEVICE":        args.device,
        "EPOCHS":        args.epochs,
        "BATCH_SIZE":    args.batch,
        "PATIENCE":      args.patience,
        "WORKERS":       args.workers,
        "CACHE":         args.cache,
        "OPTIMIZER":     args.optimizer,
        "LR0":           args.lr0,
        "LRF":           args.lrf,
        "MOMENTUM":      args.momentum,
        "WEIGHT_DECAY":  args.weight_decay,
        "WARMUP_EPOCHS": args.warmup_epochs,
        "AMP":           args.amp,
        "HSV_H":         args.hsv_h,
        "HSV_S":         args.hsv_s,
        "HSV_V":         args.hsv_v,
        "MOSAIC":        args.mosaic,
        "MIXUP":         args.mixup,
        "FLIPUD":        args.flipud,
        "FLIPLR":        args.fliplr,
        "DEGREES":       args.degrees,
        "TRANSLATE":     args.translate,
        "SCALE":         args.scale,
        "PERSPECTIVE":   args.perspective,
        "ERASING":       args.erasing,
        "BETA":          args.beta,
        "PROJECT":       args.project,
        "NAME":          args.name,
        "OUTPUT_XLSX":   args.output,
    }

    train(
        data=args.data,
        class_counts=args.counts,
        cfg=override,
    )