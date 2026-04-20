from openpyxl import load_workbook
from openpyxl.chart import LineChart, BarChart, Reference
from openpyxl.chart.label import DataLabelList

XLSX_PATH = "runs/faster_cnn/fasterrcnn_mobilenet_training.xlsx"
METRICS_SHEET = "Training Metrics"
CM_SHEET = "Confusion Matrix"
CHARTS_SHEET = "Charts"

HEADER_ROW = 2
DATA_START = 3


def build_line_chart(ws, title, y_axis, series_specs, data_end):
    chart = LineChart()
    chart.title = title
    chart.y_axis.title = y_axis
    chart.x_axis.title = "Epoch"
    chart.height = 9
    chart.width = 17
    chart.style = 2

    for col, label in series_specs:
        values = Reference(ws, min_col=col, min_row=HEADER_ROW, max_col=col, max_row=data_end)
        chart.add_data(values, titles_from_data=True)

    epochs = Reference(ws, min_col=1, min_row=DATA_START, max_row=data_end)
    chart.set_categories(epochs)
    return chart


def build_cm_chart(cm_ws):
    chart = BarChart()
    chart.type = "col"
    chart.style = 11
    chart.title = "Confusion Matrix (Counts)"
    chart.y_axis.title = "Count"
    chart.x_axis.title = "Actual Class"
    chart.height = 9
    chart.width = 17

    data = Reference(cm_ws, min_col=3, min_row=7, max_col=6, max_row=10)
    chart.add_data(data, titles_from_data=True, from_rows=False)
    cats = Reference(cm_ws, min_col=3, min_row=8, max_row=10)
    chart.set_categories(cats)
    chart.dataLabels = DataLabelList(showVal=True)
    return chart


def main():
    wb = load_workbook(XLSX_PATH)
    ws = wb[METRICS_SHEET]

    data_end = ws.max_row
    while data_end >= DATA_START and ws.cell(row=data_end, column=1).value in (None, ""):
        data_end -= 1

    if CHARTS_SHEET in wb.sheetnames:
        del wb[CHARTS_SHEET]
    charts_ws = wb.create_sheet(CHARTS_SHEET)

    charts = [
        build_line_chart(ws, "Box Loss (Train vs Val)", "Box Loss",
                         [(3, "Train"), (4, "Val")], data_end),
        build_line_chart(ws, "Class Loss (Train vs Val)", "Class Loss",
                         [(5, "Train"), (6, "Val")], data_end),
        build_line_chart(ws, "DFL Loss (Train vs Val)", "DFL Loss",
                         [(7, "Train"), (8, "Val")], data_end),
        build_line_chart(ws, "Precision (Val)", "Precision",
                         [(9, "Val")], data_end),
        build_line_chart(ws, "Recall (Val)", "Recall",
                         [(11, "Val")], data_end),
        build_line_chart(ws, "F1 Score (Val)", "F1 Score",
                         [(13, "Val")], data_end),
        build_line_chart(ws, "mAP@0.5 & mAP@0.5:0.95", "mAP",
                         [(15, "mAP@0.5"), (16, "mAP@0.5:0.95")], data_end),
        build_line_chart(ws, "Accuracy (Val)", "Accuracy",
                         [(17, "Val")], data_end),
        build_line_chart(ws, "Learning Rate", "LR",
                         [(2, "Learning Rate")], data_end),
    ]

    anchors = ["B2", "L2", "B20", "L20", "B38", "L38", "B56", "L56", "B74"]
    for anchor, chart in zip(anchors, charts):
        charts_ws.add_chart(chart, anchor)

    cm_ws = wb[CM_SHEET]
    charts_ws.add_chart(build_cm_chart(cm_ws), "B92")

    wb.save(XLSX_PATH)
    print(f"Charts written to {XLSX_PATH} (sheet: {CHARTS_SHEET}), epochs: {data_end - DATA_START + 1}")


if __name__ == "__main__":
    main()
