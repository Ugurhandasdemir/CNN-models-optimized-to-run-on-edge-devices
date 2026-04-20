"""
training_logger.py — Model Egitim Metriklerini Excel'e Kaydeden Logger
======================================================================

Herhangi bir object detection framework'u ile kullanilabilir:
  - HuggingFace Trainer (RT-DETR, DETR, vb.)
  - Ultralytics YOLO
  - Custom PyTorch training loop

KULLANIM:
    from training_logger import TrainingLogger

    logger = TrainingLogger(
        output_path="results/yolov8_experiment.xlsx",
        model_name="YOLOv8n-thermal",
    )

    # Her epoch sonunda:
    logger.log_epoch(
        epoch=0,
        lr=0.01,
        box_loss_train=1.8,
        box_loss_val=1.85,
        cls_loss_train=3.4,
        cls_loss_val=3.3,
        precision_val=0.64,
        recall_val=0.68,
        map50=0.28,
        map50_95=0.1,
    )

    # Egitim bittiginde confusion matrix kaydet:
    logger.log_confusion_matrix(
        matrix=[[85, 5, 0], [3, 90, 2], [1, 4, 70]],
        class_names=["Person", "Car", "OtherVehicle"],
    )

    logger.save()
"""

import shutil
from pathlib import Path
from typing import Optional

import numpy as np

try:
    import openpyxl
    from openpyxl.utils import get_column_letter
except ImportError:
    raise ImportError("openpyxl gerekli: pip install openpyxl")


TEMPLATE_PATH = Path(__file__).parent / "model_training_template.xlsx"

# Excel sutun haritalamasi (1-indexed)
_COL = {
    "epoch": 1,            # A
    "lr": 2,               # B
    "box_loss_train": 3,   # C
    "box_loss_val": 4,     # D
    "cls_loss_train": 5,   # E
    "cls_loss_val": 6,     # F
    "dfl_loss_train": 7,   # G
    "dfl_loss_val": 8,     # H
    "precision_val": 9,    # I
    "precision_train": 10, # J
    "recall_val": 11,      # K
    "recall_train": 12,    # L
    "f1_val": 13,          # M (formul)
    "f1_train": 14,        # N
    "map50": 15,           # O
    "map50_95": 16,        # P
    "accuracy_val": 17,    # Q
}

# Veri satirlari 3. satirdan baslar (1-2 baslik)
_DATA_START_ROW = 3


class TrainingLogger:
    """
    Excel tabanli egitim metrik logger'i.

    Template dosyasindan kopyalayarak yeni bir Excel olusturur
    ve her epoch sonunda metrikleri yazar.

    Parametreler
    ------------
    output_path : str | Path
        Cikti Excel dosyasinin yolu.

    model_name : str, varsayilan ""
        Model adi. Confusion Matrix sheet basliginda kullanilir.

    template_path : str | Path | None, varsayilan None
        Ozel template dosyasi. None ise varsayilan template kullanilir.
    """

    def __init__(
        self,
        output_path: str | Path,
        model_name: str = "",
        template_path: str | Path | None = None,
    ):
        self.output_path = Path(output_path)
        self.model_name = model_name
        self._current_row = _DATA_START_ROW

        # Template'i kopyala
        src = Path(template_path) if template_path else TEMPLATE_PATH
        if not src.exists():
            raise FileNotFoundError(f"Template bulunamadi: {src}")

        self.output_path.parent.mkdir(parents=True, exist_ok=True)
        shutil.copy2(src, self.output_path)

        self.wb = openpyxl.load_workbook(self.output_path)
        self.ws_metrics = self.wb["Training Metrics"]
        self.ws_cm = self.wb["Confusion Matrix"]

        # Ornek veriyi temizle (3. satirdan itibaren)
        self._clear_sample_data()

        # Model adini yaz
        if model_name:
            self.ws_cm["B2"] = f"Confusion Matrix \u2014 {model_name}"

    def _clear_sample_data(self):
        """Template'teki ornek verileri temizler."""
        for row in range(_DATA_START_ROW, self.ws_metrics.max_row + 1):
            for col in range(1, len(_COL) + 1):
                self.ws_metrics.cell(row=row, column=col).value = None

    def log_epoch(
        self,
        epoch: int,
        lr: Optional[float] = None,
        box_loss_train: Optional[float] = None,
        box_loss_val: Optional[float] = None,
        cls_loss_train: Optional[float] = None,
        cls_loss_val: Optional[float] = None,
        dfl_loss_train: Optional[float] = None,
        dfl_loss_val: Optional[float] = None,
        precision_val: Optional[float] = None,
        precision_train: Optional[float] = None,
        recall_val: Optional[float] = None,
        recall_train: Optional[float] = None,
        f1_train: Optional[float] = None,
        map50: Optional[float] = None,
        map50_95: Optional[float] = None,
        accuracy_val: Optional[float] = None,
    ):
        """
        Bir epoch'un metriklerini Excel'e yazar.

        Sadece mevcut metrikleri gecin, None olanlar bos birakilir.
        F1 (val) sutunu otomatik formul olarak yazilir.

        Parametreler
        ------------
        epoch : int
            Epoch numarasi (0-indexed).
        lr : float, optional
            Learning rate.
        box_loss_train, box_loss_val : float, optional
            Box/regression loss (train ve val).
        cls_loss_train, cls_loss_val : float, optional
            Classification loss (train ve val).
        dfl_loss_train, dfl_loss_val : float, optional
            Distribution focal loss (train ve val). YOLO'ya ozel.
        precision_val, precision_train : float, optional
            Precision metrigi.
        recall_val, recall_train : float, optional
            Recall metrigi.
        f1_train : float, optional
            F1 score (train). Val icin formul otomatik yazilir.
        map50 : float, optional
            mAP@0.5 (IoU=0.5).
        map50_95 : float, optional
            mAP@0.5:0.95 (COCO metrigi).
        accuracy_val : float, optional
            Validation accuracy.
        """
        row = self._current_row
        ws = self.ws_metrics

        metrics = {
            "epoch": epoch,
            "lr": lr,
            "box_loss_train": box_loss_train,
            "box_loss_val": box_loss_val,
            "cls_loss_train": cls_loss_train,
            "cls_loss_val": cls_loss_val,
            "dfl_loss_train": dfl_loss_train,
            "dfl_loss_val": dfl_loss_val,
            "precision_val": precision_val,
            "precision_train": precision_train,
            "recall_val": recall_val,
            "recall_train": recall_train,
            "f1_train": f1_train,
            "map50": map50,
            "map50_95": map50_95,
            "accuracy_val": accuracy_val,
        }

        for key, value in metrics.items():
            if value is not None:
                col = _COL[key]
                ws.cell(row=row, column=col, value=round(value, 6))

        # F1 (val) formulu: =IF((I{row}+K{row})=0, 0, 2*I*K/(I+K))
        f1_col = _COL["f1_val"]
        prec_col_letter = get_column_letter(_COL["precision_val"])
        rec_col_letter = get_column_letter(_COL["recall_val"])
        ws.cell(row=row, column=f1_col).value = (
            f"=IF(({prec_col_letter}{row}+{rec_col_letter}{row})=0,0,"
            f"2*{prec_col_letter}{row}*{rec_col_letter}{row}"
            f"/({prec_col_letter}{row}+{rec_col_letter}{row}))"
        )

        self._current_row += 1

    def log_confusion_matrix(
        self,
        matrix: list[list[int]] | np.ndarray,
        class_names: list[str],
    ):
        """
        Confusion matrix'i Excel'e yazar.

        Parametreler
        ------------
        matrix : list[list[int]] veya np.ndarray, shape=(C, C)
            Confusion matrix. satir=gercek sinif, sutun=tahmin.

        class_names : list[str]
            Sinif isimleri. Ornek: ["Person", "Car", "OtherVehicle"]
        """
        ws = self.ws_cm
        matrix = np.array(matrix)
        n = len(class_names)

        # Onceki veriyi temizle (7. satirdan itibaren)
        for row in range(7, ws.max_row + 5):
            for col in range(1, n + 10):
                ws.cell(row=row, column=col).value = None

        # Basliklar
        header_row = 7
        ws.cell(row=header_row, column=3, value="Predicted \u2192")
        for j, name in enumerate(class_names):
            ws.cell(row=header_row, column=4 + j, value=name)

        ws.cell(row=header_row + 1, column=2, value="Actual \u2193")

        # Ham degerler
        for i, name in enumerate(class_names):
            data_row = header_row + 1 + i
            ws.cell(row=data_row, column=3, value=name)
            for j in range(n):
                ws.cell(row=data_row, column=4 + j, value=int(matrix[i, j]))

        # Normalized confusion matrix (%)
        norm_start = header_row + n + 3
        ws.cell(row=norm_start, column=3, value="Normalized Confusion Matrix (%)")

        ws.cell(row=norm_start + 1, column=3, value="Actual \u2193")
        for j, name in enumerate(class_names):
            ws.cell(row=norm_start + 1, column=4 + j, value=name)

        for i, name in enumerate(class_names):
            norm_row = norm_start + 2 + i
            ws.cell(row=norm_row, column=3, value=name)
            for j in range(n):
                # Formul: ham_deger / satir_toplami
                first_col = get_column_letter(4)
                last_col = get_column_letter(4 + n - 1)
                raw_row = header_row + 1 + i
                cell_ref = f"{get_column_letter(4 + j)}{raw_row}"
                sum_ref = f"SUM({first_col}{raw_row}:{last_col}{raw_row})"
                ws.cell(row=norm_row, column=4 + j).value = (
                    f"=IF({sum_ref}=0,0,{cell_ref}/{sum_ref})"
                )

    def save(self):
        """Excel dosyasini kaydeder."""
        self.wb.save(self.output_path)
        print(f"Metrikler kaydedildi: {self.output_path}")

    def close(self):
        """Workbook'u kapatir."""
        self.wb.close()

    def __enter__(self):
        return self

    def __exit__(self, *args):
        self.save()
        self.close()


# ──────────────────────────────────────────────────────────────
# HuggingFace Trainer Callback
# ──────────────────────────────────────────────────────────────

try:
    from transformers import TrainerCallback

    class ExcelLoggerCallback(TrainerCallback):
        """
        HuggingFace Trainer ile kullanilabilen Excel logger callback.

        Kullanim:
            logger = TrainingLogger("results/rtdetr.xlsx", model_name="RT-DETR-r18")

            trainer = Trainer(
                ...,
                callbacks=[ExcelLoggerCallback(logger)],
            )
            trainer.train()
            logger.save()
        """

        def __init__(self, logger: TrainingLogger):
            self.logger = logger

        def on_evaluate(self, args, state, control, metrics=None, **kwargs):
            if metrics is None:
                return

            epoch = int(state.epoch) if state.epoch else 0
            lr = state.learning_rate if hasattr(state, "learning_rate") else None

            # HuggingFace metrics key'lerini maple
            self.logger.log_epoch(
                epoch=epoch,
                lr=lr,
                box_loss_train=metrics.get("train_loss"),
                box_loss_val=metrics.get("eval_loss"),
                cls_loss_val=metrics.get("eval_cls_loss"),
                precision_val=metrics.get("eval_precision"),
                recall_val=metrics.get("eval_recall"),
                map50=metrics.get("eval_map_50", metrics.get("eval_map")),
                map50_95=metrics.get("eval_map"),
            )

except ImportError:
    pass
